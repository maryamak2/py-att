import openpyxl
import tkinter as tk
import sys
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

fromdata = openpyxl.load_workbook('Attendance.xlsx') #to import the attendance excel file.

window = tk.Tk()
window.geometry("400x200")
window.title("Attendance")
label1 = ttk.Label(window, text="Please select the year :", font=("Times New Roman", 12)).grid(column=0, row=5, padx=10,pady=30)
label2 = ttk.Label(window, text="Please select the week number :", font=("Times New Roman", 12)).grid(column=0, row=10,padx=10, pady=30)
year = ('first', 'second', 'third', 'forth')
lec_number = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15')


def checkcmbo1(): #A function to choose & return which year is taking the lecture.
    if years.get() == "first":
        file_name = 'EECE_1st_year.xlsx'
        return file_name
    elif years.get() == "second":
        file_name = 'EECE_2nd_year.xlsx'
        return file_name
    elif years.get() == "third":
        file_name = 'EECE_3rd_year.xlsx'
        return file_name
    elif years.get() == "forth":
        file_name = 'EECE_4th_year.xlsx'
        return file_name
    else:
        messagebox.showwarning("warning", "this file is not existed")
        sys.exit()


def checkcmbo2(): #A function to determine the index of the column according to the week we are in.
    for g in range(16):                  # 16 count for weeks 1 to 15
        if number.get() == str(g):    
            value = g + 4
            return value  
    else:
        messagebox.showwarning("warning", "this week is not existed") #if entered any number out of the range a warning appears.
        sys.exit()


years = tk.StringVar()  
number = tk.StringVar()
chosenhall = ttk.Combobox(window, width=20, textvariable=years) #links the variable years to the current value of the combobox.
chosenhall['values'] = year
chosenlec = ttk.Combobox(window, width=20, textvariable=number) #links the variable number to the current value of the combobox.
chosenlec['values'] = lec_number

chosenhall.grid(column=1, row=5)           ##determine the position of the combobox in the window.
chosenhall.current()
chosenlec.grid(column=1, row=10)
chosenlec.current()
window.mainloop()

data = checkcmbo1()  #to get the year chosen.
k = checkcmbo2()     #to get the index of the column in which "absent" or "present" is written
todata = openpyxl.load_workbook(data)  #to import the basic excel file that contains list of student ids.

fromsheet = fromdata.active        #import the sheet the has the form responses.
tosheet = todata.active            #import the basic sheet the has all student ids.

for i in fromsheet.iter_rows():        #loop to iterate over rows of attendance sheet to get each id and search for it in the basic sheet.
    studentcode = i[1].value           #store each id -existed in column B "student code"-  in variable named studentcode.
    for j in tosheet.iter_rows():      #iterate over rows of basic sheet of student ids to search for the current id.
        rownum = j[3].row              #store the index of the current row into variable named rownum.
        if j[3].value == studentcode:  #compare the current id with each id in basic sheet.
            print(j[3].value)          #if the 2 ids are identical, print the id.
            tosheet.cell(row=rownum, column=k).value = i[5].value #transfer the data in the attendance sheet in the column of attendance to the basic sheet in the specic row and column"week".

# aloop to print absent if student didn't indeed submit in the form.
char = chr(ord('A') + k - 1)              # determine the index of the column according to the chosen week
torow = char + str(tosheet.max_row)       #to get the order of the maximum occupied row (last student) in excel sheet.
fromrow = char + str(2)                    #to get the row (student) from which we start the loop.
for column in tosheet[fromrow:torow]:     #iterate over rows in the specified column.
    for index, cell in enumerate(column): #access the column's element from its index.
        if cell.value != "absent" and cell.value != "present":   # if there is no value in attendance column, print absent.
            new_data = ["absent"]         
            cell.value = new_data[index]

todata.save(data)


