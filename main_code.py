import openpyxl
import sys
import haversine as hs
import pandas as pd
import tkinter as tk
from tkinter import*
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
from haversine import Unit

book=load_workbook('Attendance.xlsx')          #to import the attendance excel file.
sheet=book.active                              #to import the active sheet in the excel file.
     
#create a tkinter window having a combobox to choose the hall.   
window = tk.Tk()
window.geometry("400x200")
window.title("Attendance")
label1 = ttk.Label( window, text = "Lecture Hall", foreground = "black", font = ("Times New Roman", 16)).grid(row = 0, column = 1)
label2 = ttk.Label( window, text = "Please select the Lecture Hall :", font = ("Times New Roman", 12)).grid(column = 0, row = 5, padx = 10, pady = 30)
hall_names = ('8208','8209','8309','8310','trial')

#A function to assign the coordinates of the origin (the center) of each hall then return the origin of the chosen one.
def checkcmbo1():

    if halls.get() == "8208":
       origin = (30.0256375,31.2113679) 
       return origin 
    elif halls.get() == "8209":
        origin = (30.0258096,31.211353) 
        return origin 
    elif halls.get() == "8309":
        origin = (30.0257751,31.2113461) 
        return origin 
    elif halls.get() == "8310":
       origin = (30.0258805,31.211384) 
       return origin 
    elif halls.get() == "trial":                 #a trial option to do expeiments.
       origin = (29.9490924,30.9330594) 
       return origin 
    else:
       messagebox.showwarning("warning","this hall is not existed")
       sys.exit()
    
 #A function to calculate & return the distance between the origin and the corner of chosen hall, to determine the range of each hall.   
def checkcmbo2():

    if halls.get() == "8208":       
       origin = (30.0256375,31.2113679)                     #assign the geocode coordinates (latitude & longitude) of the origin.
       destination = (30.0257546,31.2113306)                #assign the geocode coordinates (latitude & longitude) of the corner.
       hs.haversine(origin,destination)                     #calculate the distance bet origin & corner.
       Distance=hs.haversine(origin,destination,Unit.METERS)    #to get the distance in meters.
       return  Distance
    elif halls.get() == "8209":
        origin = (30.0258096,31.211353) 
        destination = (30.0257024,31.2113552) 
        hs.haversine(origin,destination)
        Distance=hs.haversine(origin,destination,Unit.METERS)
        return  Distance
    elif halls.get() == "8309":
        origin = (30.0257751,31.2113461) 
        destination = (30.0257984,31.211367) 
        hs.haversine(origin,destination)
        Distance=hs.haversine(origin,destination,Unit.METERS)
        return  Distance 
    elif halls.get() == "8310":
        origin = (30.0258805,31.211384) 
        destination = (30.025854,31.2113474)
        hs.haversine(origin,destination)
        Distance=hs.haversine(origin,destination,Unit.METERS)
        return  Distance 
    elif halls.get() == "trial":
        origin = (29.9490924,30.9330594) 
        destination = (29.949063,30.9330334)
        hs.haversine(origin,destination)
        Distance=hs.haversine(origin,destination,Unit.METERS)
        return  Distance 
    else:
       messagebox.showwarning("warning","this hall is not existed")
       sys.exit()


halls = tk.StringVar()                              #to manage the value of the widget.
chosenhall = ttk.Combobox ( window, width = 20, textvariable = halls )     #links the variable halls to the current value of the combobox.
chosenhall['values'] = hall_names                                          #choosing the hall depending on its name.

    
chosenhall.grid(column = 1, row = 5)                      #determine the position of the combobox in the window.
chosenhall.current()
window.mainloop()

            
filedata1  =pd.read_excel("Attendance.xlsx")              #read the excel file in a variable named filedata1.
longitude=filedata1['longitude'].values.tolist()          #convert the longitude column in excel into list named longitude.
latitude=filedata1['latitude'].values.tolist()            #convert the latitude column in excel into list named latitude.

count=0                                                   #counetr to go throuh all elements in latitude and longitude lists.
origin =  checkcmbo1()                                    #get the origin coordinates of the chosen hall
main_distance = checkcmbo2()                              #get the distance bet origin & corner of chosen hall in meters.

norows = "f" + str(sheet.max_row)                         #to get the order of the maximum occupied row in excel sheet.
for column in sheet['f2':norows]:                         #iterate over all rows of column f "attendance" in excel sheet.
    for index,cell in enumerate(column):                  #access the column's element from its index.
        destination = (latitude[count], longitude[count] )         #get the geocode coordinates of each student.
        hs.haversine(origin,destination)                           #calculate the distance bet origin of chosen hall & submitted location of student
        Distance=hs.haversine(origin,destination,Unit.METERS)      #get the distance in meters in a variable named distance.
        if Distance<main_distance:                                 #check location of student with respect to range of the hall.
          new_data=["present"]
          cell.value=new_data[index]                               #if student is existed within the range of the hall, print present.
        else:
          new_data=["absent"]
          cell.value=new_data[index] 
        print("distance in m : ",Distance)                         #if student is not existed within the range of the hall, print absent.
        count=count+1
          
book.save('Attendance.xlsx')